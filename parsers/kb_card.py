import csv
from datetime import date
from pathlib import Path
from models import Transaction

DATE_COL = "이용일"
MERCHANT_COL = "이용하신곳"
AMOUNT_COL = "국내이용금액(원)"  # XLS 헤더의 \n은 파싱 시 제거


def parse(filepath: str, encoding: str = "euc-kr") -> list[Transaction]:
    ext = Path(filepath).suffix.lower()
    if ext == ".xls":
        return _parse_xls(filepath)
    if ext == ".xlsx":
        return _parse_xlsx(filepath)
    return _parse_csv(filepath, encoding)


def _parse_csv(filepath: str, encoding: str) -> list[Transaction]:
    transactions = []
    with open(filepath, encoding=encoding) as f:
        reader = csv.DictReader(f)
        for row in reader:
            txn = _make_transaction(row)
            if txn:
                transactions.append(txn)
    return transactions


def _parse_xls(filepath: str) -> list[Transaction]:
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    # 헤더 행 탐색 (DATE_COL이 있는 첫 번째 행)
    header_idx = next(
        (i for i in range(ws.nrows)
         if DATE_COL in [str(c).strip().replace("\n", "") for c in ws.row_values(i)]),
        None,
    )
    if header_idx is None:
        raise ValueError(f"'{DATE_COL}' 컬럼을 찾을 수 없습니다. DATE_COL 상수를 실제 컬럼명으로 수정해주세요.")

    headers = [str(c).strip().replace("\n", "") for c in ws.row_values(header_idx)]
    transactions = []
    for i in range(header_idx + 1, ws.nrows):
        row_values = ws.row_values(i)
        if not any(c for c in row_values if c):
            continue
        row_dict = {headers[j]: str(row_values[j]).strip() if j < len(row_values) else "" for j in range(len(headers))}
        txn = _make_transaction(row_dict)
        if txn:
            transactions.append(txn)
    return transactions


def _parse_xlsx(filepath: str) -> list[Transaction]:
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx = next(
        (i for i, row in enumerate(all_rows)
         if row and DATE_COL in [str(c).strip().replace("\n", "") for c in row if c is not None]),
        None,
    )
    if header_idx is None:
        raise ValueError(f"'{DATE_COL}' 컬럼을 찾을 수 없습니다.")

    headers = [str(c).strip().replace("\n", "") if c is not None else "" for c in all_rows[header_idx]]
    transactions = []
    for row in all_rows[header_idx + 1:]:
        if not row or not any(c for c in row if c is not None):
            continue
        row_dict = {
            headers[j]: (str(row[j]).strip() if j < len(row) and row[j] is not None else "")
            for j in range(len(headers))
        }
        txn = _make_transaction(row_dict)
        if txn:
            transactions.append(txn)
    return transactions


def _make_transaction(row: dict) -> Transaction | None:
    date_str = row.get(DATE_COL, "").strip()
    amount_str = row.get(AMOUNT_COL, "").strip()
    merchant = row.get(MERCHANT_COL, "").strip()

    if not date_str or not amount_str or not merchant:
        return None

    amount = int(float(amount_str.replace(",", "")))
    if amount == 0:  # 국내 금액이 0 = 해외 전용 거래, 스킵
        return None

    return Transaction(
        date=_parse_date(date_str),
        amount=amount,
        merchant=merchant,
        category="",
        source="kb_card",
    )


def _parse_date(value: str) -> date:
    try:
        return date.fromisoformat(value)  # YYYY-MM-DD
    except ValueError:
        year, month, day = value.split(".")
        return date(int(year), int(month), int(day))
